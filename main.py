import os
import openpyxl as xl
from collections import defaultdict
from info import getISBNRetailData, retailerInformation

def main():
    print("Welcome to the python tool which helps you sell your textbooks quickly for maximum profit!💰🪙 💴")
    isbnData = input("Please enter a list of comma separated isbns or a filepath to an Excel sheet with the isbns in the first column:\n")
    usingExcelSheet = any(letter for letter in isbnData if letter in set("abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ"))
    isbns = []
    if usingExcelSheet:
        # Assume its a path to an Excel file
        excel_filepath = isbnData.strip('"').strip("'")    
        wb = xl.open(excel_filepath)
        for sheetName in ["Input", "Sheet", "Sheet1"]:
            try:
                sheet = wb[sheetName]
                break
            except Exception:
                pass
        else:
            sheet = wb.active
        row = 1
        while sheet.cell(row=row, column=1).value:
            isbns.append(sheet.cell(row=row, column=1).value)
            row += 1
    else:
        # Assume its a comma separated list
        isbns = [int(isbn.strip()) for isbn in isbnData.split(",")]


    # Group each book with its best selling website
    retailData = {isbn: getISBNRetailData(isbn) for isbn in isbns}
    sellingGroups = defaultdict(list)
    for isbn in isbns:
        data = retailData[isbn]
        sellingGroups[data[0]["retailer"]].append(data[0])


    # Create the output excel document
    if not usingExcelSheet:
        wb = xl.Workbook()
        excel_filepath = input("Please enter the filepath you'd like to save your output to:\n")
        if excel_filepath[-5:] != ".xlsx":
            excel_filepath = os.path.join(excel_filepath, "bookdealer.xlsx")
        wb.save(filename=excel_filepath)
        
    try:
        wb.remove(wb["Output"])
    except Exception:
        pass
    finally:
        wb.create_sheet("Output")
        sheet = wb["Output"]


    # Retrieve or create the integer and currency styles
    if "integer_style" not in [style for style in wb.named_styles]:
        integer_style = xl.styles.NamedStyle(name="integer_style", number_format="0")
    else:
        # Retrieve the existing style
        integer_style = next(style for style in wb.named_styles if style == "integer_style")

    if "currency_style" not in [style for style in wb.named_styles]:
        currency_style = xl.styles.NamedStyle(name="currency_style", number_format='"$"#,##0.00')
    else:
        # Retrieve the existing style
        currency_style = next(style for style in wb.named_styles if style == "currency_style")
    

    center = xl.styles.Alignment(horizontal="center")
    bold = xl.styles.Font(bold=True)
    red_fill = xl.styles.PatternFill(start_color="FF3F3F", end_color="FF3F3F", fill_type="solid")
    orange_fill = xl.styles.PatternFill(start_color="F0904E", end_color="F0904E", fill_type="solid")
    green_fill = xl.styles.PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")

    grandTotal = 0
    openColumn = 1
    groupingWidth = 3

    for retailer, booksToSell in sorted(sorted(sellingGroups.items(), reverse=True, key=lambda e: sum(b["price"] for b in e[1])), reverse=True, key=lambda e: retailerInformation[e[0]]["minimumOrder"] <= sum(b["price"] for b in e[1])):
        openRow = 6
        sheet.merge_cells(start_column=openColumn, start_row=openRow, end_column=openColumn+groupingWidth-1, end_row=openRow)

        # Column Dimensions
        sheet.column_dimensions[xl.utils.get_column_letter(openColumn)].width = 20
        sheet.column_dimensions[xl.utils.get_column_letter(openColumn+1)].width = 15
        sheet.column_dimensions[xl.utils.get_column_letter(openColumn+2)].width = 10

        # Retailer title
        sheet.cell(row=openRow, column=openColumn).value = retailer
        if retailer == "Worthless Doorstops/Paperweights": sheet.cell(row=openRow, column=openColumn).fill = orange_fill
        #sheet.cell(row=openRow, column=openColumn).hyperlink = 
        sheet.cell(row=openRow, column=openColumn).alignment = center
        sheet.cell(row=openRow, column=openColumn).font = xl.styles.Font(bold=True, size=12)
        openRow += 1

        # Column Titles
        sheet.cell(row=openRow, column=openColumn).value = "Book Title"
        sheet.cell(row=openRow, column=openColumn).alignment = center
        sheet.cell(row=openRow, column=openColumn).font = bold
        sheet.cell(row=openRow, column=openColumn+1).value = "ISBN"
        sheet.cell(row=openRow, column=openColumn+1).alignment = center
        sheet.cell(row=openRow, column=openColumn+1).font = bold
        sheet.cell(row=openRow, column=openColumn+2).value = "Price"
        sheet.cell(row=openRow, column=openColumn+2).alignment = center
        sheet.cell(row=openRow, column=openColumn+2).font = bold
        openRow += 1
        
        total = 0
        for data in booksToSell:
            # Book Title Cell
            sheet.cell(row=openRow, column=openColumn).value = data["title"]
            sheet.cell(row=openRow, column=openColumn).hyperlink = f"https://bookscouter.com/book/{data['isbn']}-{data['slug']}?type=sell"
            

            # ISBN Cell
            sheet.cell(row=openRow, column=openColumn+1).style = integer_style
            sheet.cell(row=openRow, column=openColumn+1).value = data["isbn"]
            #sheet.cell(row=openRow, column=openColumn+1).hyperlink = data["imageURL"]

            # Price Cell
            sheet.cell(row=openRow, column=openColumn+2).style = currency_style
            sheet.cell(row=openRow, column=openColumn+2).value = data["price"]
            sheet.cell(row=openRow, column=openColumn+2).hyperlink = data["retailerURL"]
            total += data["price"]

            openRow += 1

        # Total Row
        openRow = max(13, openRow)
        sheet.cell(row=openRow, column=openColumn).value = "Total:"
        sheet.cell(row=openRow, column=openColumn).font = bold

        sheet.cell(row=openRow, column=openColumn+2).style = currency_style
        sheet.cell(row=openRow, column=openColumn+2).font = bold
        sheet.cell(row=openRow, column=openColumn+2).value = total

        openRow += 1

        # Danger Row
        if retailerInformation[retailer]["minimumOrder"] > total:
            sheet.cell(row=openRow, column=openColumn).fill = red_fill
            sheet.cell(row=openRow, column=openColumn).font = bold
            sheet.cell(row=openRow, column=openColumn).value = "Minimum Valid Order:"

            sheet.cell(row=openRow, column=openColumn+1).fill = red_fill

            sheet.cell(row=openRow, column=openColumn+2).style = currency_style
            sheet.cell(row=openRow, column=openColumn+2).fill = red_fill
            sheet.cell(row=openRow, column=openColumn+2).font = bold
            sheet.cell(row=openRow, column=openColumn+2).value = retailerInformation[retailer]["minimumOrder"]
            openRow += 1

        openColumn += groupingWidth + 1

        grandTotal += total

    # Grand Total
    sheet.merge_cells(start_column=1, start_row=1, end_column=groupingWidth, end_row=2)
    sheet.cell(row=1, column=1).font = xl.styles.Font(bold=True, size=24, name="French Script MT")
    sheet.cell(row=1, column=1).fill = green_fill
    sheet.cell(row=1, column=1).alignment = center
    sheet.cell(row=1, column=1).value = "Grand Total"

    sheet.merge_cells(start_column=1, start_row=3, end_column=groupingWidth, end_row=4)
    sheet.cell(row=3, column=1).style = currency_style
    sheet.cell(row=3, column=1).font = xl.styles.Font(bold=True, size=24, name="French Script MT")
    sheet.cell(row=3, column=1).fill = green_fill
    sheet.cell(row=3, column=1).alignment = center
    sheet.cell(row=3, column=1).value = grandTotal

    while True:
        try:
            wb.save(filename=excel_filepath)
            break
        except PermissionError:
            input("Make sure the file is closed ❎. We got a permission error when trying to save the file. Press enter to try again.")
    print("")
    print("It's done and available in Output sheet of the excel file 📁")
    print("You're welcome 💗")



if __name__ == "__main__":
    main()
